"""
Sincroniza la colección `courses` de MongoDB con la hoja 'Facultades' del Excel.

El Excel es la fuente de verdad. El script:
  1. Lee todos los cursos del Excel (Código, Nombre, Facultad, Departamento).
  2. Los compara con los documentos actuales en la colección `courses`.
  3. Inserta los nuevos, actualiza los cambios y (opcionalmente) elimina los que ya no estén.

Uso:
  py "script/SincronizarCursos.py" --dry-run
  py "script/SincronizarCursos.py" --yes
  py "script/SincronizarCursos.py" --yes --keep-missing
"""

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from pymongo import MongoClient

MONGO_URI = "mongodb://localhost:27017"
DB_NAME = "university"
EXCEL_DEFAULT = Path(__file__).resolve().parent.parent / "Documentacion" / "Universidad.xlsx"
CURSOS_JS_DEFAULT = Path(__file__).resolve().parent.parent / "Base Datos" / "Cursos.js"
SHEET_NAME = "Facultades"


def conectar_db():
    try:
        client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        return client[DB_NAME]
    except Exception as e:
        print(f"ERROR: No se pudo conectar a MongoDB en {MONGO_URI}")
        print(f"Detalle: {e}")
        sys.exit(1)


def cargar_cursos_desde_excel(ruta_excel):
    """Parsea la hoja 'Facultades' y devuelve una lista de cursos."""
    wb = openpyxl.load_workbook(str(ruta_excel), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: No se encontró la hoja '{SHEET_NAME}' en {ruta_excel}")
        print(f"Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    max_row = ws.max_row
    max_col = ws.max_column

    # Para cada columna, encontrar la facultad/departamento más cercano hacia atrás.
    facultad_por_col = {}
    depto_por_col = {}
    for col in range(1, max_col + 1):
        for c in range(col, 0, -1):
            val = ws.cell(row=3, column=c).value
            if val is not None:
                facultad_por_col[col] = str(val).strip()
                break
        for c in range(col, 0, -1):
            val = ws.cell(row=4, column=c).value
            if val is not None:
                depto_por_col[col] = str(val).strip()
                break

    cursos = []
    for col in range(1, max_col):
        header = ws.cell(row=5, column=col).value
        if header is None or str(header).strip().lower() != "curso":
            continue

        code_col = col + 1
        facultad = facultad_por_col.get(col, "SIN_FACULTAD")
        departamento = depto_por_col.get(col, "SIN_DEPARTAMENTO")

        for row in range(6, max_row + 1):
            nombre = ws.cell(row=row, column=col).value
            codigo = ws.cell(row=row, column=code_col).value

            if nombre is None or codigo is None:
                continue

            nombre = str(nombre).strip()
            codigo = str(codigo).strip()

            if not nombre or not codigo:
                continue

            cursos.append({
                "Código": codigo,
                "Nombre": nombre,
                "Facultad": facultad,
                "Departamento": departamento,
            })

    return cursos


def detectar_duplicados_excel(cursos):
    conteo = defaultdict(list)
    for c in cursos:
        conteo[c["Código"]].append(c)
    return {k: v for k, v in conteo.items() if len(v) > 1}


def detectar_referencias_rotas(db, codigos):
    """Devuelve códigos que están referenciados en students o docentes."""
    rotas = []
    for codigo in sorted(codigos):
        en_students = db.students.count_documents({"cursos.codigo": codigo})
        en_docentes = db.docentes.count_documents({"cursos.codigo": codigo})
        if en_students or en_docentes:
            rotas.append((codigo, en_students, en_docentes))
    return rotas


def construir_cursos_finales(cursos_excel, actuales, default_credits=30):
    """Combina los datos del Excel con los créditos actuales de MongoDB."""
    cursos_finales = []
    for curso_excel in cursos_excel:
        codigo = curso_excel["Código"]
        curso_final = dict(curso_excel)
        if codigo in actuales and isinstance(actuales[codigo].get("Créditos"), (int, float)):
            curso_final["Créditos"] = actuales[codigo]["Créditos"]
        else:
            curso_final["Créditos"] = default_credits
        cursos_finales.append(curso_final)
    return cursos_finales


def escribir_cursos_js(cursos, ruta_salida):
    """Genera un archivo .js con insertMany para recrear la colección courses."""
    ruta_salida = Path(ruta_salida)

    def escapar(s):
        return str(s).replace('\\', '\\\\').replace('"', '\\"')

    lineas = [
        "show dbs",
        "",
        "use university",
        "",
        "show collections",
        "",
        "db.courses.insertMany([",
    ]

    for i, c in enumerate(cursos):
        sufijo = "," if i < len(cursos) - 1 else ""
        lineas.append("{")
        lineas.append(f'  "Código": "{escapar(c["Código"])}",')
        lineas.append(f'  "Nombre": "{escapar(c["Nombre"])}",')
        lineas.append(f'  "Facultad": "{escapar(c["Facultad"])}",')
        lineas.append(f'  "Departamento": "{escapar(c["Departamento"])}",')
        lineas.append(f'  "Créditos": {c["Créditos"]}')
        lineas.append("}" + sufijo)

    lineas.extend([
        "])",
        "",
    ])

    ruta_salida.write_text("\n".join(lineas) + "\n", encoding="utf-8")
    print(f"Archivo JS generado: {ruta_salida}")


def preparar_cambios(db, cursos_excel, delete_missing=False, default_credits=30):
    """Calcula las diferencias entre Excel y MongoDB. Retorna (a_insertar, a_actualizar, a_eliminar, actuales)."""
    excel_by_code = {c["Código"]: c for c in cursos_excel}
    actuales = {c["Código"]: c for c in db.courses.find({}, {"_id": 0})}

    a_insertar = []
    a_actualizar = []

    for codigo, curso_excel in excel_by_code.items():
        if codigo not in actuales:
            curso_nuevo = dict(curso_excel)
            curso_nuevo["Créditos"] = default_credits
            a_insertar.append(curso_nuevo)
        else:
            actual = actuales[codigo]
            cambios = {}
            for campo in ["Nombre", "Facultad", "Departamento"]:
                if actual.get(campo) != curso_excel.get(campo):
                    cambios[campo] = curso_excel[campo]
            if cambios:
                a_actualizar.append((codigo, cambios, actual, curso_excel))

    solo_mongodb = []
    a_eliminar = []
    for codigo in actuales:
        if codigo not in excel_by_code:
            solo_mongodb.append(codigo)
            if delete_missing:
                a_eliminar.append(codigo)

    return a_insertar, a_actualizar, a_eliminar, solo_mongodb, actuales


def mostrar_reporte(a_insertar, a_actualizar, a_eliminar, solo_mongodb, actuales, total_excel, db):
    print(f"\nCursos en Excel: {total_excel}")
    print(f"Cursos en MongoDB: {len(actuales)}")
    print(f"A insertar: {len(a_insertar)}")
    print(f"A actualizar: {len(a_actualizar)}")
    print(f"A eliminar: {len(a_eliminar)}")
    print(f"Solo en MongoDB: {len(solo_mongodb)}")

    if a_insertar:
        print("\n--- Cursos a insertar ---")
        for c in a_insertar:
            print(f"  {c['Código']}: {c['Nombre']} ({c['Facultad']} - {c['Departamento']})")

    if a_actualizar:
        print("\n--- Cursos a actualizar ---")
        for codigo, cambios, actual, nuevo in a_actualizar:
            print(f"  {codigo}:")
            for campo, valor_nuevo in cambios.items():
                print(f"    {campo}: '{actual.get(campo)}' -> '{valor_nuevo}'")

    if a_eliminar:
        print("\n--- Cursos a eliminar ---")
        for codigo in sorted(a_eliminar):
            actual = actuales[codigo]
            print(f"  {codigo}: {actual.get('Nombre')} ({actual.get('Facultad')} - {actual.get('Departamento')})")

        rotas = detectar_referencias_rotas(db, a_eliminar)
        if rotas:
            print("\n⚠️  ADVERTENCIA: Los siguientes códigos a eliminar tienen referencias en otras colecciones:")
            for codigo, en_students, en_docentes in rotas:
                print(f"  {codigo}: {en_students} estudiante(s), {en_docentes} docente(s)")
            print("Eliminarlos dejará referencias rotas en students/docentes.")

    if solo_mongodb:
        print("\n--- Cursos solo en MongoDB (no están en el Excel) ---")
        for codigo in sorted(solo_mongodb):
            actual = actuales[codigo]
            print(f"  {codigo}: {actual.get('Nombre')} ({actual.get('Facultad')} - {actual.get('Departamento')})")


def sincronizar(db, cursos_excel, delete_missing=False, dry_run=False, default_credits=30):
    duplicados = detectar_duplicados_excel(cursos_excel)
    if duplicados:
        print("ERROR: Hay códigos duplicados en el Excel. Corrígelos antes de continuar.")
        for codigo, items in sorted(duplicados.items()):
            for item in items:
                print(f"  {codigo}: {item['Nombre']} ({item['Facultad']} - {item['Departamento']})")
        return None

    a_insertar, a_actualizar, a_eliminar, solo_mongodb, actuales = preparar_cambios(
        db, cursos_excel, delete_missing, default_credits
    )

    mostrar_reporte(a_insertar, a_actualizar, a_eliminar, solo_mongodb, actuales, len(cursos_excel), db)

    if dry_run:
        print("\n[MODO SIMULACIÓN] No se aplicaron cambios.")

    if not (a_insertar or a_actualizar or a_eliminar):
        print("\nNo hay cambios que aplicar en MongoDB.")

    return a_insertar, a_actualizar, a_eliminar, solo_mongodb


def aplicar_cambios(db, a_insertar, a_actualizar, a_eliminar):
    if a_insertar:
        db.courses.insert_many(a_insertar)
        print(f"Insertados {len(a_insertar)} cursos.")

    for codigo, cambios, _, _ in a_actualizar:
        db.courses.update_one({"Código": codigo}, {"$set": cambios})
    if a_actualizar:
        print(f"Actualizados {len(a_actualizar)} cursos.")

    if a_eliminar:
        result = db.courses.delete_many({"Código": {"$in": a_eliminar}})
        print(f"Eliminados {result.deleted_count} cursos.")


def main():
    parser = argparse.ArgumentParser(
        description="Sincroniza la colección 'courses' de MongoDB con el Excel de facultades."
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=EXCEL_DEFAULT,
        help=f"Ruta al archivo Excel. Por defecto: {EXCEL_DEFAULT}",
    )
    parser.add_argument(
        "--delete-missing",
        action="store_true",
        dest="delete_missing",
        default=True,
        help="Elimina de MongoDB los cursos que no estén en el Excel. (Por defecto)",
    )
    parser.add_argument(
        "--keep-missing",
        action="store_false",
        dest="delete_missing",
        help="No elimina los cursos que estén en MongoDB pero no en el Excel.",
    )
    parser.add_argument(
        "--yes", "-y",
        action="store_true",
        help="Aplica los cambios sin pedir confirmación.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Muestra los cambios sin aplicarlos.",
    )
    parser.add_argument(
        "--default-credits",
        type=int,
        default=30,
        help="Créditos por defecto para cursos nuevos. Por defecto: 30.",
    )
    parser.add_argument(
        "--update-js",
        action="store_true",
        help="También regenera el archivo 'Base Datos/Cursos.js' desde el Excel.",
    )
    parser.add_argument(
        "--js-output",
        type=Path,
        default=CURSOS_JS_DEFAULT,
        help=f"Ruta del archivo .js a generar. Por defecto: {CURSOS_JS_DEFAULT}",
    )
    parser.add_argument(
        "--mongo-only",
        action="store_true",
        help="Lista los cursos que están en MongoDB pero no en el Excel, y sale.",
    )
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERROR: No se encontró el archivo Excel: {args.excel}")
        sys.exit(1)

    print(f"Leyendo cursos desde: {args.excel}")
    cursos_excel = cargar_cursos_desde_excel(args.excel)
    print(f"Se leyeron {len(cursos_excel)} cursos del Excel.")

    db = conectar_db()

    if args.mongo_only:
        excel_by_code = {c["Código"]: c for c in cursos_excel}
        actuales = {c["Código"]: c for c in db.courses.find({}, {"_id": 0})}
        solo_mongodb = [codigo for codigo in actuales if codigo not in excel_by_code]
        print(f"\nCursos exclusivos de MongoDB (no están en el Excel): {len(solo_mongodb)}")
        for codigo in sorted(solo_mongodb):
            actual = actuales[codigo]
            print(f"  {codigo}: {actual.get('Nombre')} ({actual.get('Facultad')} - {actual.get('Departamento')})")
        return

    resultado = sincronizar(
        db, cursos_excel, args.delete_missing, args.dry_run, args.default_credits
    )

    if resultado is None:
        sys.exit(1)

    a_insertar, a_actualizar, a_eliminar, solo_mongodb = resultado
    hay_cambios_db = bool(a_insertar or a_actualizar or a_eliminar)

    if args.dry_run:
        return

    if hay_cambios_db:
        if not args.yes:
            respuesta = input("\n¿Aplicar los cambios en MongoDB? [s/N]: ").strip().lower()
            if respuesta not in ("s", "si", "sí"):
                print("Operación cancelada.")
                return
        aplicar_cambios(db, a_insertar, a_actualizar, a_eliminar)
        print("Sincronización con MongoDB completada.")

    if args.update_js:
        actuales = {c["Código"]: c for c in db.courses.find({}, {"_id": 0})}
        cursos_finales = construir_cursos_finales(cursos_excel, actuales, args.default_credits)
        escribir_cursos_js(cursos_finales, args.js_output)
        print("Archivo Cursos.js regenerado desde el Excel.")


if __name__ == "__main__":
    main()
